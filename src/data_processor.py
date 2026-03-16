import pandas as pd
import ruptures as rpt
from openpyxl import load_workbook

# =========================
# 2023–2025 mmcfd sheet
# =========================
def read_mmcfd_2023_25(file_path, sheet_name="2023-25 - mmcfd Mpi3j"):
    """
    Read 2023–2025 marketable gas production from Excel (mmcfd) and return tidy DataFrame
    """
    PROVINCES = [
        "Nova Scotia","New Brunswick","Ontario","Saskatchewan",
        "Alberta","British Columbia","NWT & Yukon","Canada Total"
    ]

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Read D10–K45
    data_rows = []
    for row in ws.iter_rows(
        min_row=10, max_row=45,
        min_col=4, max_col=11,
        values_only=True
    ):
        data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=PROVINCES)

    # Create Year / Month
    df["Year"] = [2023]*12 + [2024]*12 + [2025]*12
    df["Month_num"] = list(range(1, 13)) * 3

    # Create Date
    df["Date"] = pd.to_datetime(
    dict(year=df["Year"], month=df["Month_num"], day=1)
    )
    
    # Convert to %B
    df["Month"] = df["Date"].dt.strftime("%B")

    df = df.drop(columns=["Month_num", "Date"])

    # tidy
    tidy = df.melt(
        id_vars=["Year", "Month"],
        var_name="Province",
        value_name="Production_mmcfd"
    )

    tidy["Production_mmcfd"] = (
        pd.to_numeric(tidy["Production_mmcfd"], errors="coerce")
        .round(0)
        .astype("Int64")
    )

    MONTH_ORDER = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
    ]
    
    tidy["Month"] = pd.Categorical(
        tidy["Month"],
        categories=MONTH_ORDER,
        ordered=True
        )

    return tidy.sort_values(
        ["Year", "Month", "Province"]
    ).reset_index(drop=True)


# =========================
# 2000–2025 mmcfd sheet
# =========================
def read_mmcfd_2000_25(file_path, sheet_name="2000+ - mmcfd Mpi3j"):
    """
    Read 2000–2025 marketable gas production from Excel (mmcfd) and return tidy DataFrame
    """
    PROVINCE_MAP = {
        "NS": "Nova Scotia",
        "NB": "New Brunswick",
        "Ontario": "Ontario",
        "Sask": "Saskatchewan",
        "Alberta": "Alberta",
        "BC": "British Columbia",
        "Territories": "NWT & Yukon",
        "Total Canada": "Canada Total"
    }

    COLUMN_ORDER = [
        ("C", "NS"),
        ("D", "NB"),
        ("E", "Ontario"),
        ("F", "Sask"),
        ("G", "Alberta"),
        ("H", "BC"),
        ("I", "Territories"),
        ("K", "Total Canada"),
    ]

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    records = []
    for row in range(12, 324):
        date_cell = ws[f"A{row}"].value
        if date_cell is None:
            continue

        date = pd.to_datetime(date_cell)
        year = date.year
        month = date.strftime("%B")

        for col_letter, prov_key in COLUMN_ORDER:
            value = ws[f"{col_letter}{row}"].value
            records.append({
                "Year": year,
                "Month": month,
                "Province": PROVINCE_MAP[prov_key],
                "Production_mmcfd": value
            })

    df = pd.DataFrame(records)

    df["Production_mmcfd"] = (
        pd.to_numeric(df["Production_mmcfd"], errors="coerce")
        .round(0)
        .astype("Int64")
    )

    MONTH_ORDER = [
        "January","February","March","April","May","June",
        "July","August","September","October","November","December"
    ]
    df["Month"] = pd.Categorical(df["Month"], categories=MONTH_ORDER, ordered=True)

    return df.sort_values(
        ["Year", "Month", "Province"]
    ).reset_index(drop=True)

# =========================
# 2023–2025 e3m3d sheet
# =========================
def read_e3m3d_2023_25(file_path, sheet_name="2023-25 - 103m3d 103m3j"):
    """
    Read 2023–2025 marketable gas production from Excel (e3m3d) and return tidy DataFrame
    """
    PROVINCES = [
        "Nova Scotia","New Brunswick","Ontario","Saskatchewan",
        "Alberta","British Columbia","NWT & Yukon","Canada Total"
    ]

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Read D10–K45
    data_rows = []
    for row in ws.iter_rows(
        min_row=10, max_row=45,
        min_col=4, max_col=11,
        values_only=True
    ):
        data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=PROVINCES)

    # Create Year / Month
    df["Year"] = [2023]*12 + [2024]*12 + [2025]*12
    df["Month_num"] = list(range(1, 13)) * 3

    # Create Date
    df["Date"] = pd.to_datetime(
    dict(year=df["Year"], month=df["Month_num"], day=1)
    )
    
    # Convert to %B
    df["Month"] = df["Date"].dt.strftime("%B")

    df = df.drop(columns=["Month_num", "Date"])

    # tidy
    tidy = df.melt(
        id_vars=["Year", "Month"],
        var_name="Province",
        value_name="Production_e3m3d"
    )

    tidy["Production_e3m3d"] = (
        pd.to_numeric(tidy["Production_e3m3d"], errors="coerce")
        .round(0)
        .astype("Int64")
    )

    MONTH_ORDER = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
    ]
    
    tidy["Month"] = pd.Categorical(
        tidy["Month"],
        categories=MONTH_ORDER,
        ordered=True
        )

    return tidy.sort_values(
        ["Year", "Month", "Province"]
    ).reset_index(drop=True)


# =========================
# 2000–2025 e3m3d sheet
# =========================
def read_e3m3d_2000_25(file_path, sheet_name="2000+ - 103m3d 103m3j"):
    """
    Read 2000–2025 marketable gas production from Excel (e3m3d) and return tidy DataFrame
    """
    PROVINCE_MAP = {
        "NS": "Nova Scotia",
        "NB": "New Brunswick",
        "Ontario": "Ontario",
        "Sask": "Saskatchewan",
        "Alberta": "Alberta",
        "BC": "British Columbia",
        "Territories": "NWT & Yukon",
        "Total Canada": "Canada Total"
    }

    COLUMN_ORDER = [
        ("C", "NS"),
        ("D", "NB"),
        ("E", "Ontario"),
        ("F", "Sask"),
        ("G", "Alberta"),
        ("H", "BC"),
        ("I", "Territories"),
        ("K", "Total Canada"),
    ]

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    records = []
    for row in range(12, 324):
        date_cell = ws[f"A{row}"].value
        if date_cell is None:
            continue

        date = pd.to_datetime(date_cell)
        year = date.year
        month = date.strftime("%B")

        for col_letter, prov_key in COLUMN_ORDER:
            value = ws[f"{col_letter}{row}"].value
            records.append({
                "Year": year,
                "Month": month,
                "Province": PROVINCE_MAP[prov_key],
                "Production_e3m3d": value
            })

    df = pd.DataFrame(records)

    df["Production_e3m3d"] = (
        pd.to_numeric(df["Production_e3m3d"], errors="coerce")
        .round(0)
        .astype("Int64")
    )

    MONTH_ORDER = [
        "January","February","March","April","May","June",
        "July","August","September","October","November","December"
    ]
    df["Month"] = pd.Categorical(df["Month"], categories=MONTH_ORDER, ordered=True)

    return df.sort_values(
        ["Year", "Month", "Province"]
    ).reset_index(drop=True)

# ================================================================================

def aggregate_annual(df, col="Production_e3m3d"):
    df = df.groupby("Year")[col].mean().reset_index()
    return df

def add_rolling(df, col="Production_e3m3d", window=3):
    df = df.copy()
    df[f"{col}_roll{window}"] = df[col].rolling(window=window, center=True).mean()
    return df

def detect_change_points(df, col="Production_e3m3d", pen=1):
    """
    Returns detected change years
    """
    series = df[col].dropna().values
    years = df["Year"].dropna().values
    algo = rpt.Pelt(model="rbf").fit(series)
    breakpoints = algo.predict(pen=pen)
    # Remove the last "break" (end) option
    if breakpoints[-1] == len(years):
        breakpoints = breakpoints[:-1]
    change_years = [years[bp-1] for bp in breakpoints if bp > 0]
    return change_years
